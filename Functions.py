import pandas as pd
import sys
import os
import openpyxl
from openpyxl.styles import Font,Alignment
def adjust_excel_format(filename):
    # 加载Excel文件
    workbook = openpyxl.load_workbook(filename)

    # 设置字体
    font = Font(name='Times New Roman', size=12)

    for sheet in workbook.sheetnames:
        worksheet = workbook[sheet]
        
        # 设置单元格居中
        alignment = Alignment(horizontal='center', vertical='center')
        
        # 遍历所有单元格，设置字体
        for row in worksheet.iter_rows():
            for cell in row:
                cell.font = font
                cell.alignment = alignment

        # 自动调整列宽
        for column in worksheet.columns:
            max_length = 0
            column = [cell for cell in column if cell.value]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[openpyxl.utils.get_column_letter(column[0].column)].width = adjusted_width

    # 保存修改后的文件
    workbook.save(filename)